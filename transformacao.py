import pandas as pd
import os
import datetime
from openpyxl.utils import get_column_letter
import openpyxl 

# Lista com os nomes exatos dos arquivos Excel que serão unificados (verificar como o Script irá encontrar futuramente os arquivos, pois eles vem com nomes diferentes e um ID de Identificação)
arquivos = [
    'planilhas/Pontomais-Colaborador.xlsx', 
    'planilhas/Pontomais-Ponto.xlsx', 
    'planilhas/Pontomais-Turnos.xlsx'
]
# Lista com os nomes que serão usados como o cabeçalho principal para cada arquivo.
nomes_relatorios = [
    'Relatório Colaborador', 
    'Relatório Ponto', 
    'Relatório Turnos'
]

# Lista de tuplas para especificar quais colunas remover
colunas_para_excluir = [
    ('Relatório Colaborador', 'Cargo'),
    ('Relatório Colaborador', 'E-mail'),
    ('Relatório Colaborador', 'CPF'),
    ('Relatório Colaborador', 'Centro de custo'),
    ('Relatório Colaborador', 'Data de admissão'),
    ('Relatório Ponto', 'Unnamed: 0'),
    ('Relatório Ponto', 'Unnamed: 1'),
    ('Relatório Ponto', 'Unnamed: 2'),
    ('Relatório Turnos', 'Código'),
    ('Relatório Turnos', 'Descrição'),
    ('Relatório Turnos', 'Virada do turno'),
    ('Relatório Turnos', 'Tipo'),
    ('Relatório Turnos', 'Limite de horas extras'),
    ('Relatório Turnos', 'Dia') 
]

linhas_pular = 2      # Número de linhas a serem ignoradas no início de cada arquivo
indice_cabecalho = 0  # A linha que será considerada o cabeçalho após pular as linhas acima

data_atual = datetime.datetime.now()
timestamp_diario = data_atual.strftime("%Y-%m-%d") # Formato de saída do nome do arquivo: ANO-MÊS-DIA
arquivo_final = f'Pontomais-unificado_{timestamp_diario}.xlsx'

def unificar_e_formatar_excel(lista_arquivos, chaves_cabecalho, skip_rows, header_row_index, arquivo_saida, colunas_a_remover):
    lista_de_dados = [] # Lista vazia para armazenar os dados de cada planilha 
    print("Iniciando o processo...")

    # Validação inicial para garantir que as listas de configuração correspondem
    if len(lista_arquivos) != len(chaves_cabecalho):
        print("Erro: O número de arquivos não é igual ao número de nomes de relatórios.")
        return # Encerra a função se houver erro

    # LER CADA ARQUIVO EXCEL
    for arquivo in lista_arquivos:
        # Verifica se o arquivo realmente existe no diretório antes de tentar lê-lo
        if os.path.exists(arquivo):
            try:
                # Tenta ler o arquivo Excel 
                df = pd.read_excel(arquivo, skiprows=skip_rows, header=header_row_index)
                # Adiciona o DataFrame lido à lista de dados
                lista_de_dados.append(df)
                print(f"Arquivo '{arquivo}' lido com sucesso...")
            except Exception as e:
                print(f"Erro ao ler o arquivo '{arquivo}': {e}")
        else:
            print(f"Atenção: O arquivo '{arquivo}' não foi encontrado e será ignorado.")

    if not lista_de_dados:
        print("\nNenhum dado foi lido. O programa será encerrado.")
        return

    # UNIFICAR E LIMPAR OS DADOS
    print("\nJuntando os dados com cabeçalhos de múltiplos níveis...")
    # Concatena todos os DataFrames da lista
    # axis=1: unifica as planilhas lado a lado 
    # keys: cria um cabeçalho de primeiro nível usando os nomes dos relatórios
    dados_unificados = pd.concat(lista_de_dados, axis=1, keys=chaves_cabecalho)
    
    print(f"Removendo {len(colunas_a_remover)} coluna(s) especificadas...")
    # Remove as colunas listadas na configuração
    # errors='ignore': evita que o script quebre se uma coluna a ser removida não for encontrada
    dados_unificados = dados_unificados.drop(columns=colunas_a_remover, errors='ignore')
    print("Colunas removidas com sucesso.")

    # SALVAR E FORMATAR O ARQUIVO FINAL 
    print(f"\nSalvando e formatando o arquivo final: '{arquivo_saida}'...")
    try:
        # pd.ExcelWriter permite salvar e ao mesmo tempo acessar o objeto da planilha para formatação
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
            # Salva o DataFrame unificado em uma aba chamada 'Dados Unificados', sem a coluna de índice do pandas
            dados_unificados.to_excel(writer, sheet_name='Dados Unificados')

            # Obtém o objeto da planilha para poder manipulá-lo
            worksheet = writer.sheets['Dados Unificados']

            # Loop para autoajuste da largura das colunas
            for i, column in enumerate(dados_unificados.columns):
                # Calcula a largura necessária: o maior valor entre o comprimento do cabeçalho de nível 1,
                # o cabeçalho de nível 2 e o maior dado da coluna, mais um espaço extra (+2)
                header_level_1 = len(str(column[0]))
                header_level_2 = len(str(column[1]))
                max_len_data = dados_unificados[column].astype(str).map(len).max()
                column_width = max(header_level_1, header_level_2, max_len_data) + 2
                
                # Converte o índice numérico da coluna (0, 1, 2...) para a letra correspondente no Excel (A, B, C...)
                # Adiciona +2 porque a primeira coluna (idx=1) é a do índice que será removida, e a contagem começa em 1
                column_letter = get_column_letter(i + 2)
                
                # Aplica a largura calculada à coluna
                worksheet.column_dimensions[column_letter].width = column_width

        print("Arquivo salvo inicialmente. Agora, removendo a coluna de índice...")
        workbook = openpyxl.load_workbook(arquivo_saida)  # O to_excel salva uma coluna de índice por padrão, removido com openpyxl
        sheet = workbook.active
        sheet.delete_cols(idx=1) # Deleta a primeira coluna 
        workbook.save(arquivo_saida) # Salva o arquivo com a alteração

        print(f"\n Sucesso! Arquivo '{arquivo_saida}' finalizado e formatado corretamente.")
    except Exception as e:
        print(f"\n Erro ao salvar ou formatar o arquivo final: {e}")

if __name__ == "__main__":
    # Chama a função principal e passa todas as variáveis de configuração
    unificar_e_formatar_excel(
        lista_arquivos=arquivos,
        chaves_cabecalho=nomes_relatorios,
        skip_rows=linhas_pular,
        header_row_index=indice_cabecalho, 
        arquivo_saida=arquivo_final,
        colunas_a_remover=colunas_para_excluir
    )