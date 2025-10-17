import pandas as pd
import os
import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font
import glob

diretorio_planilhas = '../planilhas'
diretorio_relatorio = '../relatorio'

os.makedirs(diretorio_relatorio, exist_ok=True)

linhas_pular = 2
linhas_pular_ponto = 3
indice_cabecalho = 0

colunas_a_remover = {
    'colaboradores': [
        'PIS', 'Cargo', 'CPF', 'E-mail', 'Centro de custo', 'Data de admissão'
    ],
    'turnos': [
        'Limite de horas extras', 'Tipo', '1ª Saída prevista', '2ª Saída prevista', 'Virada do turno'
    ],
    'registros_de_ponto': [
        'Equipe', 'Turno'
    ]
}

data_atual = datetime.datetime.now()
timestamp_diario = data_atual.strftime("%Y-%m-%d")
nome_arquivo = f'Pontomais-unificado_{timestamp_diario}.xlsx'
arquivo_saida_final = os.path.join(diretorio_relatorio, nome_arquivo)

# Obtem arquivos mais recentes 
def obter_arquivos(diretorio, quantidade):
    print(f"Buscando os {quantidade} arquivos mais recentes em '{diretorio}'...")
    caminho_completo = os.path.join(diretorio, '*.xlsx')
    lista_de_arquivos = glob.glob(caminho_completo)

    if len(lista_de_arquivos) < quantidade:
        print(f"Atenção: Foram encontrados apenas {len(lista_de_arquivos)} arquivos. São necessários {quantidade}.")
        return []

    lista_de_arquivos.sort(key=os.path.getmtime)
    return lista_de_arquivos[:quantidade]

# Gera os nomes dos relatórios
def nomes_relatorios(lista_arquivos):
    nomes = []
    for arquivo in lista_arquivos:
        nome_base = os.path.basename(arquivo)
        nome_limpo = nome_base.replace('.xlsx', '')
        nomes.append(f'Relatório {nome_limpo}')
    return nomes

# Unifica as planilhas
def unificar_planilhas(lista_arquivos, chaves_cabecalho, header_row_index):
    lista_de_dados = []
    print("\nIniciando o processo de unificação...")

    for arquivo in lista_arquivos:
        try:
            nome_arquivo_base = os.path.basename(arquivo).lower()

            if 'registros_de_ponto' in nome_arquivo_base:
                print(f"   - Arquivo '{os.path.basename(arquivo)}' (Relatório de Ponto) -> Pulando {linhas_pular_ponto} linhas.")
                df = pd.read_excel(arquivo, skiprows=linhas_pular_ponto, header=header_row_index, engine='openpyxl')
            else:
                print(f"   - Arquivo '{os.path.basename(arquivo)}' -> Pulando {linhas_pular} linhas.")
                df = pd.read_excel(arquivo, skiprows=linhas_pular, header=header_row_index, engine='openpyxl')

            lista_de_dados.append(df)

        except Exception as e:
            print(f"Erro ao ler o arquivo '{arquivo}': {e}")
            return None

    if not lista_de_dados:
        print("\nNenhum dado foi lido. O programa será encerrado.")
        return None

    print("\nJuntando os dados com cabeçalhos de múltiplos níveis...")
    dados_unificados = pd.concat(lista_de_dados, axis=1, keys=chaves_cabecalho)

    print("\nRemovendo colunas desnecessárias com base em palavras-chave...")
    colunas_a_remover_final = []

    for col_tupla in dados_unificados.columns:
        report_name = str(col_tupla[0]).lower()
        column_name = str(col_tupla[1])

        for keyword, column_list in colunas_a_remover.items():
            if keyword in report_name and column_name in column_list:
                colunas_a_remover_final.append(col_tupla)
                break

    if colunas_a_remover_final:
        dados_unificados = dados_unificados.drop(columns=colunas_a_remover_final)
        print(f"   - {len(colunas_a_remover_final)} coluna(s) removida(s) com sucesso.")
    else:
        print("   - Nenhuma coluna para remover foi encontrada.")

    return dados_unificados
    
def colaboradores_faltantes(dados_unificados, arquivos, linhas_pular=2, linhas_pular_ponto=3):
    print("\nIniciando comparação entre Colaboradores e Registros de Ponto...")

    arquivo_colaboradores = None
    arquivo_ponto = None

    for arq in arquivos:
        nome = os.path.basename(arq).lower()
        if 'colaboradores' in nome:
            arquivo_colaboradores = arq
        elif 'registros_de_ponto' in nome:
            arquivo_ponto = arq

    if not arquivo_colaboradores or not arquivo_ponto:
        print("ERRO: Não foi possível localizar os arquivos de colaboradores ou de registros de ponto.")
        return dados_unificados, pd.DataFrame()

    df_colab = pd.read_excel(arquivo_colaboradores, skiprows=linhas_pular, engine='openpyxl')
    df_ponto = pd.read_excel(arquivo_ponto, skiprows=linhas_pular_ponto, engine='openpyxl')

    df_colab['Nome'] = df_colab['Nome'].astype(str).str.strip().str.lower()
    df_ponto['Nome'] = df_ponto['Nome'].astype(str).str.strip().str.lower()

    nomes_colab = set(df_colab['Nome'])
    nomes_ponto = set(df_ponto['Nome'])
    faltantes = nomes_colab - nomes_ponto

    df_faltantes = df_colab[df_colab['Nome'].isin(faltantes)][['Nome', 'Equipe']].copy()

    if 'Turno' in df_colab.columns:
        df_turno = df_colab[['Nome', 'Turno']].copy()
        df_faltantes = pd.merge(df_faltantes, df_turno, on='Nome', how='left')
    else:
        df_faltantes['Turno'] = '—'

    try:
        caminho_turnos = os.path.join('../planilhas', 'turnos_extraidos.xlsx')
        if os.path.exists(caminho_turnos):
            df_turnos = pd.read_excel(caminho_turnos, engine='openpyxl')
            df_turnos['descrição'] = df_turnos['descrição'].astype(str).str.lower().str.strip()

            df_faltantes['Código do Turno'] = '—'
            df_faltantes['1ª Entrada prevista'] = '—'

            for i, row in df_faltantes.iterrows():
                turno_colab = str(row['Turno']).lower().strip()
                codigo_encontrado = ''
                entrada_encontrado = ''
                
                for _, turno_ref in df_turnos.iterrows():
                    desc_turno = str(turno_ref['descrição']).lower().strip()
                    if desc_turno[:12] in turno_colab or turno_colab[:12] in desc_turno:
                        codigo_encontrado = turno_ref.get('codigo', '—')
                        entrada_encontrado = turno_ref.get('primeira_entrada_prevista', '—')
                        break  
                
                df_faltantes.at[i, 'Código do Turno'] = codigo_encontrado if codigo_encontrado else '—'
                df_faltantes.at[i, '1ª Entrada prevista'] = entrada_encontrado if entrada_encontrado else '—'

            print("Códigos de turno e primeira entrada associados com sucesso.")
        else:
            print("Atenção: Arquivo 'turnos_extraidos.xlsx' não encontrado na pasta '../planilhas'.")
    except Exception as e:
        print(f"Erro ao associar código de turno e entrada: {e}")

    return dados_unificados, df_faltantes

def salvar_arquivo_excel(dados_unificados, df_faltantes, arquivo_saida):
    print(f"\nSalvando e formatando o arquivo final: '{arquivo_saida}'...")

    if isinstance(dados_unificados.columns, pd.MultiIndex):
        dados_unificados.columns = [' - '.join([str(x) for x in col if x]).strip() for col in dados_unificados.columns]

    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
        dados_unificados.to_excel(writer, sheet_name='Dados Unificados', index=False)
        if not df_faltantes.empty:
            df_faltantes.to_excel(writer, sheet_name='Faltantes', index=False)

    # Pós-formatação
    wb = load_workbook(arquivo_saida)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(10, max_length + 2)
    wb.save(arquivo_saida)
    print("Arquivo salvo e formatado com sucesso.")


def limpar_arquivos_originais(lista_arquivos):
    print("\nIniciando limpeza dos arquivos de origem...")
    for arquivo in lista_arquivos:
        try:
            os.remove(arquivo)
            print(f"   - Arquivo '{os.path.basename(arquivo)}' removido com sucesso.")
        except OSError as e:
            print(f"   - Erro ao remover o arquivo '{os.path.basename(arquivo)}': {e}")

# Execução
if __name__ == "__main__":
    arquivos_para_processar = obter_arquivos(diretorio_planilhas, 3)

    if arquivos_para_processar:
        print("\nArquivos encontrados (do mais antigo para o mais novo):")
        for f in arquivos_para_processar:
            print(f"   - {os.path.basename(f)}")

        chaves_relatorios = nomes_relatorios(arquivos_para_processar)
        dados_unificados = unificar_planilhas(
            lista_arquivos=arquivos_para_processar,
            chaves_cabecalho=chaves_relatorios,
            header_row_index=indice_cabecalho
        )

        if dados_unificados is not None:
            dados_unificados, df_faltantes = colaboradores_faltantes(dados_unificados, arquivos_para_processar)
            salvar_arquivo_excel(dados_unificados, df_faltantes, arquivo_saida_final)
            limpar_arquivos_originais(arquivos_para_processar)
    else:
        print("\nProcesso não iniciado devido à falta de arquivos.")
